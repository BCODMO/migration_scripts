#!/usr/bin/env python3
"""
Script to detect potential Excel formatting bugs in pipeline specifications.
Parses pipeline-spec.yaml files and checks Excel files for number format issues.

Requirements:
- convert.py module must be present with the following functions:
  - handle_general_old(number_format, value)
  - handle_general_new(number_format, value)
  - convert_old(number_format, value)
  - convert_new(number_format, value)

Output includes:
- Pipeline spec name
- Pipeline name
- Excel file path
- Sheet name where error occurred
- Error location (row/column)
- Old and new values from conversion
- Format type (General or Other)
- Last updated timestamp of the pipeline spec

The script checks ALL sheets in each Excel file and records bugs for each sheet
where they occur, allowing comprehensive bug detection across workbooks.
"""

import fnmatch
from urllib.parse import unquote

import boto3
import yaml
import openpyxl
from openpyxl.utils import get_column_letter
import json
import os
import tempfile
from typing import List, Dict, Any, Tuple, Optional
from urllib.parse import urlparse
import logging
from convert import handle_general_old, handle_general_new, convert_old, convert_new

# Constants
BUCKET_NAME = "laminar-dump"  # Change this to your actual bucket name
OUTPUT_FILE = "potentially_buggy_pipelines.json"
LOG_FILE = "excel_bug_detection.log"

# Configure logging
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# Create formatters and add it to handlers
log_format = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")

# Console handler (stdout)
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_handler.setFormatter(log_format)

# File handler
file_handler = logging.FileHandler(LOG_FILE)
file_handler.setLevel(logging.INFO)
file_handler.setFormatter(log_format)

# Add handlers to the logger
logger.addHandler(console_handler)
logger.addHandler(file_handler)


# S3 client
s3_client = boto3.client("s3")


def list_pipeline_specs(bucket: str) -> List[str]:
    """
    List all pipeline-spec.yaml files in the bucket.
    """
    pipeline_specs = []
    paginator = s3_client.get_paginator("list_objects_v2")

    for page in paginator.paginate(Bucket=bucket):
        if "Contents" in page:
            for obj in page["Contents"]:
                if obj["Key"].endswith("pipeline-spec.yaml"):
                    pipeline_specs.append(obj["Key"])

    return pipeline_specs


def download_s3_file(s3_path: str) -> Optional[str]:
    """
    Download a file from S3 to a temporary location.
    Returns the local file path or None if download fails.
    """
    try:
        parsed = urlparse(s3_path)
        bucket = parsed.netloc
        key = parsed.path.lstrip("/")

        # Get file size
        response = s3_client.head_object(Bucket=bucket, Key=key)
        file_size = response["ContentLength"]

        # Format file size for readable display
        if file_size < 1024 * 1024:  # Less than 1 MB
            size_str = f"{file_size:,} bytes"
        elif file_size < 1024 * 1024 * 1024:  # Less than 1 GB
            size_str = f"{file_size / (1024 * 1024):.2f} MB"
        else:  # 1 GB or more
            size_str = f"{file_size / (1024 * 1024 * 1024):.2f} GB"

        logger.info(f"Processing file: {s3_path} (Size: {size_str})")

        # Download to temp file
        temp_file = tempfile.NamedTemporaryFile(
            delete=False, suffix=os.path.splitext(key)[1]
        )
        s3_client.download_file(bucket, key, temp_file.name)
        temp_file.close()

        return temp_file.name
    except Exception as e:
        logger.error(f"Failed to download {s3_path}: {e}")
        return None


def parse_pipeline_spec(bucket: str, key: str) -> Tuple[Dict[str, Any], Optional[str]]:
    """
    Download and parse a pipeline-spec.yaml file.
    Returns (parsed_yaml, last_modified_date) tuple.
    """
    try:
        response = s3_client.get_object(Bucket=bucket, Key=key)
        content = response["Body"].read().decode("utf-8")
        last_modified = response["LastModified"].strftime("%Y-%m-%d %H:%M:%S UTC")
        return yaml.safe_load(content), last_modified
    except Exception as e:
        logger.error(f"Failed to parse {key}: {e}")
        return {}, None


def extract_excel_files(pipeline_spec: Dict[str, Any]) -> List[Tuple[str, str]]:
    """
    Extract Excel file paths from pipeline specification.
    Returns list of (pipeline_name, excel_path) tuples.
    """
    excel_files = []

    # Get the first key (pipeline name)
    if not pipeline_spec:
        return excel_files

    pipeline_name = list(pipeline_spec.keys())[0]
    pipeline_data = pipeline_spec[pipeline_name]

    if "pipeline" not in pipeline_data:
        return excel_files

    for step in pipeline_data.get("pipeline", []):
        if step.get("run") == "bcodmo_pipeline_processors.load":
            params = step.get("parameters", {})
            if params.get("format") == "xlsx":
                from_paths = params.get("from", [])
                if params.get("input_path_pattern", False):
                    new_from_list = []
                    for p in params.get("from", []):
                        temp_from_list = []
                        if p.startswith("s3://"):
                            # Handle s3 pattern
                            try:
                                bucket, path = p[5:].split("/", 1)
                            except ValueError:
                                raise Exception(
                                    f"Improperly formed S3 url passed to the load step: {p}"
                                )

                            path_parts = path.split("/", 1)
                            object_id = ""
                            if len(path_parts) > 0:
                                object_id = path_parts[0]
                            s3 = boto3.resource("s3")

                            bucket_obj = s3.Bucket(bucket)

                            matches = fnmatch.filter(
                                [
                                    unquote(obj.key)
                                    for obj in bucket_obj.objects.filter(
                                        Prefix=object_id
                                    ).all()
                                ],
                                path,
                            )
                            for match in matches:
                                temp_from_list.append(f"s3://{bucket}/{match}")

                            if not len(temp_from_list):
                                raise Exception(
                                    f"No objects found in S3 matching the glob pattern {p}. Are you sure the files have been properly staged?"
                                )
                        else:
                            raise Exception("Found a pipeline with non s3 start")
                        new_from_list += temp_from_list

                    from_paths = new_from_list

                if isinstance(from_paths, str):
                    from_paths = [from_paths]
                for path in from_paths:
                    excel_files.append((pipeline_name, path.strip()))

    return excel_files


def check_excel_for_bug(excel_path: str) -> List[Tuple[str, int, str, Any, Any, str]]:
    """
    Check if an Excel file is potentially affected by the bug.
    Returns list of (sheet_name, row, column, old_value, new_value, format_type) for all bugs detected.
    """
    bugs_found = []
    local_path = None
    try:
        # Download the file if it's from S3
        if excel_path.startswith("s3://"):
            local_path = download_s3_file(excel_path)
            if not local_path:
                return bugs_found
        else:
            local_path = excel_path

        # Load the workbook
        wb = openpyxl.load_workbook(local_path, data_only=True)

        for sheet_name in wb.sheetnames:
            logger.info(f"  Checking sheet: {sheet_name}")
            sheet = wb[sheet_name]
            sheet_has_bug = False

            subsuquent_all_none = 0
            for row_idx, row in enumerate(sheet.iter_rows(), start=1):
                all_none = True
                for col_idx, cell in enumerate(row, start=1):
                    if cell.value is not None:
                        all_none = False
                    # Check if cell has numeric value
                    if isinstance(cell.value, (int, float)):
                        format_str = cell.number_format or "General"

                        if format_str == "General":
                            # Handle General format
                            old_result = handle_general_old(format_str, cell.value)
                            new_result = handle_general_new(format_str, cell.value)
                            format_type = "General"
                        else:
                            # Handle other formats
                            old_result = convert_old(format_str, cell.value)
                            new_result = convert_new(format_str, cell.value)
                            format_type = "Other"

                        if old_result != new_result:
                            col_letter = get_column_letter(col_idx)
                            logger.warning(
                                f"    Bug detected in sheet '{sheet_name}' at {col_letter}{row_idx} (Format: {format_type}, {format_str})"
                            )
                            bugs_found.append(
                                (
                                    sheet_name,
                                    row_idx,
                                    col_letter,
                                    old_result,
                                    new_result,
                                    format_type,
                                )
                            )
                            sheet_has_bug = True
                            break  # Move to next sheet after finding first bug in this sheet

                if sheet_has_bug:
                    break  # Break out of row loop to move to next sheet
                if all_none:
                    subsuquent_all_none += 1

                if subsuquent_all_none >= 100:
                    break

        wb.close()

    except Exception as e:
        logger.error(f"Error checking Excel file {excel_path}: {e}")
    finally:
        # Clean up temp file
        if local_path and excel_path.startswith("s3://") and os.path.exists(local_path):
            os.unlink(local_path)

    return bugs_found


def main():
    """
    Main execution function.
    """
    logger.info(
        f"Starting Excel bug detection script - Logging to console and {LOG_FILE}"
    )
    buggy_pipelines = []
    total_excel_files_processed = 0

    # Test mode - uncomment to test with specific file
    # test_mode = True
    test_mode = False

    if test_mode:
        # Test with specific pipeline spec
        pipeline_specs = ["962507_TEST/1/data/pipeline-spec.yaml"]
        pipeline_specs = ["930172/1/data/pipeline-spec.yaml"]
        pipeline_specs = ["749429/2/data/pipeline-spec.yaml"]
        logger.info("Running in TEST MODE with specific pipeline spec")
    else:
        # Get all pipeline specs
        logger.info(f"Scanning bucket: {BUCKET_NAME}")
        pipeline_specs = list_pipeline_specs(BUCKET_NAME)
        logger.info(f"Found {len(pipeline_specs)} pipeline-spec.yaml files")

    # Process each pipeline spec
    for idx, spec_key in enumerate(pipeline_specs, 1):
        progress_pct = (idx / len(pipeline_specs)) * 100
        logger.info(
            f"Processing [{idx}/{len(pipeline_specs)}] ({progress_pct:.1f}%): {spec_key}"
        )

        # Parse the pipeline spec
        pipeline_spec, last_modified = parse_pipeline_spec(BUCKET_NAME, spec_key)
        if last_modified:
            logger.info(f"  Pipeline spec last updated: {last_modified}")

        # Extract Excel files
        excel_files = extract_excel_files(pipeline_spec)
        total_excel_files_processed += len(excel_files)

        # Check each Excel file
        for pipeline_name, excel_path in excel_files:
            logger.info(f"  Checking Excel: {excel_path}")
            bugs_in_file = check_excel_for_bug(excel_path)

            if bugs_in_file:
                logger.warning(
                    f"  Found {len(bugs_in_file)} sheet(s) with potential bugs in {pipeline_name}"
                )
                for sheet_name, row, col, old_val, new_val, format_type in bugs_in_file:
                    buggy_pipelines.append(
                        {
                            "pipeline_spec": spec_key,
                            "pipeline_name": pipeline_name,
                            "excel_file": excel_path,
                            "sheet_name": sheet_name,
                            "error_location": f"{col}{row}",
                            "row": row,
                            "column": col,
                            "old_value": str(old_val),
                            "new_value": str(new_val),
                            "format_type": format_type,
                            "last_updated": last_modified or "Unknown",
                        }
                    )
                    logger.info(
                        f"    Recorded bug in sheet '{sheet_name}' (Format Type: {format_type})"
                    )

        # Log progress summary every 10 pipelines (or at the end)
        if idx % 10 == 0 or idx == len(pipeline_specs):
            remaining = len(pipeline_specs) - idx
            general_bugs = sum(
                1 for b in buggy_pipelines if b.get("format_type") == "General"
            )
            other_bugs = sum(
                1 for b in buggy_pipelines if b.get("format_type") == "Other"
            )
            unique_files = len(
                set((b["pipeline_spec"], b["excel_file"]) for b in buggy_pipelines)
            )
            logger.info(
                f"  Progress Summary: {idx} processed, {remaining} remaining, {total_excel_files_processed} Excel files checked, {len(buggy_pipelines)} buggy sheets in {unique_files} files (General: {general_bugs}, Other: {other_bugs})"
            )

    # Save results
    logger.info(f"Completed processing all {len(pipeline_specs)} pipeline specs")
    logger.info(f"Total Excel files processed: {total_excel_files_processed}")

    if total_excel_files_processed == 0:
        logger.warning("No Excel files were found to process!")

    logger.info(f"Found {len(buggy_pipelines)} potentially buggy sheets")

    # Count unique files affected
    unique_files = len(
        set((b["pipeline_spec"], b["excel_file"]) for b in buggy_pipelines)
    )
    logger.info(f"Affecting {unique_files} unique Excel file(s)")

    # Break down by format type
    general_bugs = sum(1 for b in buggy_pipelines if b.get("format_type") == "General")
    other_bugs = sum(1 for b in buggy_pipelines if b.get("format_type") == "Other")
    logger.info(f"  - General format bugs: {general_bugs}")
    logger.info(f"  - Other format bugs: {other_bugs}")

    # Sort by last updated date (most recent first) for easier prioritization
    buggy_pipelines.sort(key=lambda x: x.get("last_updated", ""), reverse=True)

    with open(OUTPUT_FILE, "w") as f:
        json.dump(buggy_pipelines, f, indent=2)

    logger.info(
        f"Results saved to {OUTPUT_FILE} (sorted by last updated date, newest first)"
    )
    logger.info(f"Complete log saved to {LOG_FILE}")

    # Also print summary
    if buggy_pipelines:
        print("\nPotentially Buggy Pipelines:")
        print("-" * 80)
        for bug in buggy_pipelines:
            print(f"Pipeline Spec: {bug['pipeline_spec']}")
            print(f"  Pipeline Name: {bug['pipeline_name']}")
            print(f"  Excel File: {bug['excel_file']}")
            print(f"  Sheet Name: {bug['sheet_name']}")
            print(
                f"  Error at: {bug['error_location']} (Row {bug['row']}, Column {bug['column']})"
            )
            print(f"  Format Type: {bug['format_type']}")
            print(f"  Old Value: {bug['old_value']}")
            print(f"  New Value: {bug['new_value']}")
            print(f"  Last Updated: {bug['last_updated']}")
            print()

        # Print summary statistics
        general_count = sum(
            1 for b in buggy_pipelines if b.get("format_type") == "General"
        )
        other_count = sum(1 for b in buggy_pipelines if b.get("format_type") == "Other")
        unique_files = len(
            set((b["pipeline_spec"], b["excel_file"]) for b in buggy_pipelines)
        )
        print("-" * 80)
        print(f"Total Excel files processed: {total_excel_files_processed}")
        print(
            f"Total bugs found: {len(buggy_pipelines)} sheets across {unique_files} file(s)"
        )
        print(f"  - General format: {general_count}")
        print(f"  - Other formats: {other_count}")

        # Show sheets affected per file
        file_sheet_counts = {}
        for bug in buggy_pipelines:
            file_key = (bug["pipeline_spec"], bug["excel_file"])
            if file_key not in file_sheet_counts:
                file_sheet_counts[file_key] = set()
            file_sheet_counts[file_key].add(bug["sheet_name"])

        if len(file_sheet_counts) <= 5:  # Only show details for small numbers of files
            print("\nSheets affected per file:")
            for (spec, excel), sheets in file_sheet_counts.items():
                print(f"  {excel}: {len(sheets)} sheet(s)")

        # Show date range of buggy pipelines
        dates = [
            b["last_updated"] for b in buggy_pipelines if b["last_updated"] != "Unknown"
        ]
        if dates:
            print(f"\nDate range of buggy pipelines:")
            print(f"  - Oldest: {min(dates)}")
            print(f"  - Newest: {max(dates)}")
    else:
        print("\n" + "=" * 80)
        print(
            f"No bugs detected in {total_excel_files_processed} Excel files across {len(pipeline_specs)} pipeline specs!"
        )
        print("=" * 80)


if __name__ == "__main__":
    main()
